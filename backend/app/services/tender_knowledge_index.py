import csv
import hashlib
import io
import re

from openpyxl import load_workbook
import xlrd

from sqlalchemy import delete,select
from sqlalchemy.orm import Session

from app.models import AuditEvent,BidDocument,TenderKnowledgeChunk
from app.services.requirement_extraction import RuleBasedRequirementExtractionProvider,SourceUnit


INDEXABLE_EXTENSIONS={"pdf","docx","txt","xlsx","xls","csv"}
INDEXER_VERSION="tender-knowledge-index-v3"


def _clause(text:str):
 match=re.match(r"\s*(?:(?:clause|section)\s+)?(\d+(?:\.\d+){1,6})\b",str(text or ""),re.I)
 return match.group(1) if match else None


def _sentences(text:str):
 parts=re.split(r"(?<=[.!?;])\s+|\n+",str(text or ""))
 return [re.sub(r"\s+"," ",x).strip() for x in parts if len(re.sub(r"\s+"," ",x).strip())>=15]


CLAUSE_LINE_RE=re.compile(r"^\s*(?:(?:clause|section)\s+)?\d+(?:\.\d+){1,6}\b",re.I)


def _clause_blocks(text:str):
 lines=[re.sub(r"\s+"," ",x).strip() for x in str(text or "").splitlines() if x.strip()]
 if not lines:return []
 blocks=[];current=[]
 for line in lines:
  if CLAUSE_LINE_RE.match(line) and current:
   blocks.append(" ".join(current));current=[line]
  else:
   current.append(line)
 if current:blocks.append(" ".join(current))
 return blocks


def _chunks(text:str,target_chars:int=1300,overlap_sentences:int=1):
 sentences=_sentences(text)
 if not sentences:return []
 result=[];current=[]
 for sentence in sentences:
  candidate=" ".join(current+[sentence])
  if current and len(candidate)>target_chars:
   result.append(" ".join(current))
   current=current[-overlap_sentences:]+[sentence] if overlap_sentences else [sentence]
  else:
   current.append(sentence)
 if current:result.append(" ".join(current))
 return result


def _row_text(values):
 return " | ".join(str(x).strip() for x in values if x not in (None,"") and str(x).strip())


def _spreadsheet_units(extension:str,content:bytes):
 units=[]
 if extension=="xlsx":
  workbook=load_workbook(io.BytesIO(content),data_only=True,read_only=True)
  for ws in workbook.worksheets:
   for row in ws.iter_rows(values_only=True):
    text=_row_text(row)
    if len(text)>=10:units.append(SourceUnit(text,section=ws.title))
 elif extension=="xls":
  workbook=xlrd.open_workbook(file_contents=content)
  for ws in workbook.sheets():
   for index in range(ws.nrows):
    text=_row_text(ws.row_values(index))
    if len(text)>=10:units.append(SourceUnit(text,section=ws.name))
 elif extension=="csv":
  rows=csv.reader(io.StringIO(content.decode("utf-8-sig",errors="replace")))
  for row in rows:
   text=_row_text(row)
   if len(text)>=10:units.append(SourceUnit(text,section="CSV"))
 return units


def index_tender_document(db:Session,document:BidDocument,storage,user_id:int,request_metadata:dict|None=None):
 if document.duplicate_of_document_id or not document.storage_path:
  raise ValueError("Document content is not available for tender indexing")
 ext=document.file_extension.lower()
 if ext not in INDEXABLE_EXTENSIONS:
  return {"document_id":document.id,"indexed":False,"chunks":0,"reason":"Document format is not text-indexable yet","version":INDEXER_VERSION}

 provider=RuleBasedRequirementExtractionProvider()
 content=storage.read(document.storage_path)
 units=_spreadsheet_units(ext,content) if ext in {"xlsx","xls","csv"} else provider.source_units(ext,content)
 db.execute(delete(TenderKnowledgeChunk).where(TenderKnowledgeChunk.source_document_id==document.id))
 created=0
 for unit in units:
  blocks=_clause_blocks(unit.text)
  if not blocks:blocks=[unit.text]
  for block in blocks:
   for text in _chunks(block):
    digest=hashlib.sha256(text.encode("utf-8")).hexdigest()
    db.add(TenderKnowledgeChunk(
     bid_project_id=document.bid_project_id,source_document_id=document.id,chunk_index=created+1,
     source_page=str(unit.page) if unit.page else None,source_clause=_clause(text),source_section=unit.section,
     text=text,text_hash=digest,word_count=len(text.split()),
     source_kind=document.document_category or document.document_type or "Tender Document",is_active=True,
    ))
    created+=1
 db.add(AuditEvent(
  user_id=user_id,bid_project_id=document.bid_project_id,event_type="tender_knowledge.indexed",
  entity_type="BidDocument",entity_id=str(document.id),request_metadata=request_metadata or {},
  details={"chunks":created,"extension":ext,"version":INDEXER_VERSION},
 ))
 db.commit()
 return {
  "document_id":document.id,"indexed":created>0,"chunks":created,
  "reason":None if created else "No extractable text was found; OCR or another extractor may be required",
  "version":INDEXER_VERSION,
 }


def tender_knowledge_status(db:Session,bid_id:int):
 docs=db.scalars(select(BidDocument).where(
  BidDocument.bid_project_id==bid_id,
  BidDocument.is_latest_revision.is_(True),
  BidDocument.document_status!="Archived",
  BidDocument.duplicate_of_document_id.is_(None),
 )).all()
 chunks=db.scalars(select(TenderKnowledgeChunk).where(
  TenderKnowledgeChunk.bid_project_id==bid_id,
  TenderKnowledgeChunk.is_active.is_(True),
 )).all()
 indexed_doc_ids={x.source_document_id for x in chunks}
 indexable=[x for x in docs if x.file_extension.lower() in INDEXABLE_EXTENSIONS]
 return {
  "summary":{
   "latest_documents":len(docs),
   "text_indexable_documents":len(indexable),
   "indexed_documents":len(indexed_doc_ids),
   "chunks":len(chunks),
   "index_coverage_percent":100.0 if not indexable else round(len({x.id for x in indexable}&indexed_doc_ids)*100/len(indexable),1),
   "needs_extraction":sum(1 for x in indexable if x.id not in indexed_doc_ids),
   "non_text_documents":sum(1 for x in docs if x.file_extension.lower() not in INDEXABLE_EXTENSIONS),
  },
  "documents":[{
   "document_id":x.id,"document_name":x.original_filename,"extension":x.file_extension,
   "indexable":x.file_extension.lower() in INDEXABLE_EXTENSIONS,
   "indexed":x.id in indexed_doc_ids,
   "chunks":sum(1 for c in chunks if c.source_document_id==x.id),
  } for x in docs],
  "version":INDEXER_VERSION,
  "note":"Index coverage describes text retrieval readiness. Scanned/image-only documents require OCR or vision before they can contribute full-text evidence.",
 }
