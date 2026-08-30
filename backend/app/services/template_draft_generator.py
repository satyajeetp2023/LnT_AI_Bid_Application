import io
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.template_population_plan import build_population_plan


ALLOWED_CHOICE_MARKS={"X","✓"}


def generate_controlled_xlsx_draft(
    db:Session,
    bid_id:int,
    template_content:bytes,
    choice_mark:str="X",
    include_suggested_text:bool=False,
    header_values:dict|None=None,
    field_overrides:dict|None=None,
):
    if choice_mark not in ALLOWED_CHOICE_MARKS:
        raise ValueError("Unsupported choice mark")
    plan=build_population_plan(db,bid_id,template_content)
    workbook=load_workbook(io.BytesIO(template_content),data_only=False)
    written=[]
    skipped=[]
    header_values=header_values or {}
    field_overrides=field_overrides or {}
    required_header_fields=[x["semantic_field"] for x in plan.get("header_inputs",[])]
    missing_header_fields=[x for x in required_header_fields if not str(header_values.get(x,"") or "").strip()]

    allowed_override_fields={}
    row_choice_keys={}
    for row in plan["rows"]:
        for field in row["fields"]:
            coordinate=field.get("coordinate")
            if not coordinate:continue
            key=f'{row["sheet"]}!{coordinate}'
            allowed_override_fields[key]={**field,"sheet":row["sheet"]}
            if field.get("semantic_field") in {"compliant_yes","compliant_no"}:
                row_choice_keys.setdefault(row["sheet"],[]).append(key)

    for sheet,keys in row_choice_keys.items():
        selected=[k for k in keys if str(field_overrides.get(k,"") or "").strip()]
        if len(selected)>1:
            raise ValueError(f"Only one compliance choice may be selected on sheet {sheet}")
    for key,value in field_overrides.items():
        field=allowed_override_fields.get(key)
        if not field:
            raise ValueError(f"Unknown template field override: {key}")
        if field.get("ownership")=="employer_only" or field.get("action")=="preserve":
            raise ValueError(f"Field {key} is not bidder-editable")
        if field.get("action") not in {"needs_review","needs_human_decision","needs_assessment","needs_input","suggest_text","leave_blank"}:
            raise ValueError(f"Field {key} cannot be manually overridden in the current workflow")

    structure=plan.get("_structure") if isinstance(plan.get("_structure"),dict) else None
    if structure is None:
        from app.services.template_structure_parser import parse_xlsx_template
        structure=parse_xlsx_template(template_content)
    for placeholder in structure.get("workbook_placeholders",[]):
        value=str(header_values.get(placeholder["semantic_field"],"") or "").strip()
        if not value:continue
        label=placeholder["label"]
        pattern=re.compile(rf"\b({re.escape(label)})\s*[-_]{{5,}}",re.I)
        for occurrence in placeholder.get("occurrences",[]):
            ws=workbook[occurrence["sheet"]]
            cell=ws[occurrence["coordinate"]]
            original=str(cell.value or "")
            updated=pattern.sub(lambda m:f"{m.group(1)}: {value}",original)
            if updated!=original:
                cell.value=updated
                written.append({"sheet":occurrence["sheet"],"coordinate":occurrence["coordinate"],"semantic_field":placeholder["semantic_field"],"value":value})

    for row in plan["rows"]:
        ws=workbook[row["sheet"]]
        for field in row["fields"]:
            coordinate=field.get("coordinate")
            if not coordinate:continue
            key=f'{row["sheet"]}!{coordinate}'
            override=field_overrides.get(key)
            if override is not None and str(override).strip():
                value=choice_mark if field.get("semantic_field") in {"compliant_yes","compliant_no"} else str(override)
                ws[coordinate]=value
                written.append({"sheet":row["sheet"],"coordinate":coordinate,"semantic_field":field.get("semantic_field"),"value":value,"source":"human_override"})
                continue
            action=field.get("action")
            semantic=field.get("semantic_field")
            proposed=field.get("proposed_value")

            if action=="propose_auto_fill":
                if semantic in {"compliant_yes","compliant_no"}:
                    if proposed:
                        ws[coordinate]=choice_mark
                        written.append({"sheet":row["sheet"],"coordinate":coordinate,"semantic_field":semantic,"value":choice_mark})
                elif proposed not in (None,""):
                    ws[coordinate]=proposed
                    written.append({"sheet":row["sheet"],"coordinate":coordinate,"semantic_field":semantic,"value":proposed})
            elif action=="suggest_text" and include_suggested_text and proposed:
                ws[coordinate]=proposed
                written.append({"sheet":row["sheet"],"coordinate":coordinate,"semantic_field":semantic,"value":proposed})
            elif action in {"needs_review","needs_human_decision","needs_assessment","needs_input"}:
                skipped.append({"sheet":row["sheet"],"coordinate":coordinate,"semantic_field":semantic,"reason":field.get("reason")})

    output=io.BytesIO()
    workbook.save(output)
    return output.getvalue(),{
        "written_fields":len(written),
        "unresolved_fields":len(skipped),
        "written":written,
        "unresolved":skipped,
        "choice_mark":choice_mark,
        "suggested_text_included":include_suggested_text,
        "header_values_applied":sorted(k for k,v in header_values.items() if str(v or "").strip()),
        "missing_header_fields":missing_header_fields,
        "missing_header_field_count":len(missing_header_fields),
        "human_overrides_applied":sum(1 for x in written if x.get("source")=="human_override"),
        "plan_version":plan.get("plan_version"),
        "generator_version":"phase5-controlled-xlsx-generator-v4",
    }
