import csv
import io
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from datetime import datetime

from docx import Document as DocxDocument
from openpyxl import load_workbook

from app.services.document_classification import extract_text
from app.services.p6_xer import parse_xer


SCHEDULE_EXTENSIONS={"xer","xml","xlsx","xls","csv","pdf","doc","docx","txt","mpp"}

ALIASES={
    "task_code":("activity id","activity code","task id","id","activity no","activity number"),
    "task_name":("activity name","activity","task name","description","activity description"),
    "wbs":("wbs","wbs name","work breakdown structure"),
    "status_code":("status","activity status"),
    "task_type":("activity type","task type","type"),
    "start_date":("start","start date","planned start","early start"),
    "finish_date":("finish","finish date","planned finish","early finish"),
    "duration":("duration","original duration","planned duration","remaining duration"),
    "total_float":("total float","float","total float hours"),
    "calendar":("calendar","calendar name"),
    "predecessors":("predecessors","predecessor","pred"),
    "resources":("resources","resource names","resource"),
}


def _norm(value):
    return re.sub(r"\s+"," ",str(value or "").strip().lower())


def _semantic(value):
    text=_norm(value)
    compact=re.sub(r"\([^)]*\)|\[[^]]*\]","",text).strip(" :*-")
    for semantic,names in ALIASES.items():
        if text in names or compact in names:return semantic
        if any(compact.startswith(name+" ") for name in names if len(name)>=5):return semantic
    return None


def _date_text(value):
    if value in (None,""):return None
    if isinstance(value,datetime):return value.isoformat()
    return str(value).strip() or None


def _number(value):
    if value in (None,""):return None
    try:return float(value)
    except (TypeError,ValueError):
        match=re.search(r"-?\d+(?:\.\d+)?",str(value))
        return float(match.group()) if match else None


def _duration_hours(value):
    if value in (None,""):return None
    if isinstance(value,(int,float)):return float(value)*8.0
    text=str(value).strip()
    iso=re.fullmatch(r"P(?:(?P<days>\d+(?:\.\d+)?)D)?(?:T(?:(?P<hours>\d+(?:\.\d+)?)H)?(?:(?P<minutes>\d+(?:\.\d+)?)M)?)?",text,re.I)
    if iso:
        return float(iso.group("days") or 0)*8+float(iso.group("hours") or 0)+float(iso.group("minutes") or 0)/60
    lower=text.lower()
    number=_number(lower)
    if number is None:return None
    if "hour" in lower:return number
    if "minute" in lower:return number/60
    if "week" in lower:return number*40
    return number*8


def _table_matrix_to_schedule(matrix:list[list],sheet_name:str=""):
    best=None
    for row_index,row in enumerate(matrix[:80]):
        mapping={}
        for col_index,value in enumerate(row):
            semantic=_semantic(value)
            if semantic and semantic not in mapping:mapping[semantic]=col_index
        if "task_name" in mapping and ("task_code" in mapping or "start_date" in mapping or "finish_date" in mapping):
            score=len(mapping)
            if best is None or score>best[0]:best=(score,row_index,mapping)
    if not best:return None
    _,header_row,mapping=best
    tasks=[];wbs_names={};rels=[]
    for index,row in enumerate(matrix[header_row+1:],1):
        def get(name):
            col=mapping.get(name)
            return row[col] if col is not None and col<len(row) else None
        name=str(get("task_name") or "").strip()
        if not name:continue
        code=str(get("task_code") or f"{sheet_name or 'SCH'}-{index:05d}").strip()
        wbs=str(get("wbs") or "").strip()
        if wbs and wbs not in wbs_names:wbs_names[wbs]=str(len(wbs_names)+1)
        task={
            "task_id":code,
            "task_code":code,
            "task_name":name,
            "wbs_id":wbs_names.get(wbs),
            "status_code":str(get("status_code") or "").strip() or None,
            "task_type":str(get("task_type") or "").strip() or "TT_Task",
            "target_start_date":_date_text(get("start_date")),
            "target_end_date":_date_text(get("finish_date")),
            "target_drtn_hr_cnt":_duration_hours(get("duration")),
            "remain_drtn_hr_cnt":_duration_hours(get("duration")),
            "total_float_hr_cnt":_number(get("total_float")),
            "clndr_id":str(get("calendar") or "").strip() or None,
            "_source_sheet":sheet_name,
        }
        tasks.append(task)
        pred_text=str(get("predecessors") or "").strip()
        if pred_text:
            for pred in re.split(r"[,;\n]+",pred_text):
                pred=pred.strip()
                if pred:
                    rels.append({"task_id":code,"pred_task_id":pred,"pred_type":"PR_FS","lag_hr_cnt":0})
    wbs=[{"wbs_id":wid,"wbs_name":name,"parent_wbs_id":None} for name,wid in wbs_names.items()]
    return {"TASK":tasks,"PROJWBS":wbs,"TASKPRED":rels}


def _spreadsheet(extension:str,content:bytes):
    candidates=[]
    if extension=="xlsx":
        wb=load_workbook(io.BytesIO(content),data_only=False,read_only=True)
        for ws in wb.worksheets:
            matrix=[list(x) for x in ws.iter_rows(values_only=True)]
            parsed=_table_matrix_to_schedule(matrix,ws.title)
            if parsed and parsed["TASK"]:candidates.append(parsed)
    elif extension=="csv":
        matrix=list(csv.reader(io.StringIO(content.decode("utf-8-sig",errors="replace"))))
        parsed=_table_matrix_to_schedule(matrix,"CSV")
        if parsed and parsed["TASK"]:candidates.append(parsed)
    if not candidates:return None
    tasks=[];wbs=[];rels=[]
    for x in candidates:
        tasks.extend(x["TASK"]);wbs.extend(x["PROJWBS"]);rels.extend(x["TASKPRED"])
    return {"PROJECT":[],"TASK":tasks,"PROJWBS":wbs,"TASKPRED":rels,"CALENDAR":[],"RSRC":[],"TASKRSRC":[]}


def _local(tag):
    return tag.rsplit("}",1)[-1]


def _child_text(node,*names):
    wanted={x.lower() for x in names}
    for child in list(node):
        if _local(child.tag).lower() in wanted:
            return (child.text or "").strip() or None
    return None


def _p6_xml(content:bytes):
    try:root=ET.fromstring(content)
    except ET.ParseError:return None
    tasks=[];rels=[];wbs=[];projects=[];resources=[];assignments=[];calendars=[]
    for node in root.iter():
        tag=_local(node.tag).lower()
        if tag in {"activity","task"}:
            code=_child_text(node,"Id","ActivityId","ObjectId","Code")
            name=_child_text(node,"Name","ActivityName")
            if not name:continue
            task_id=code or _child_text(node,"ObjectId") or str(len(tasks)+1)
            task={
                "task_id":task_id,
                "task_code":code or task_id,
                "task_name":name,
                "wbs_id":_child_text(node,"WBSObjectId","WBSId"),
                "status_code":_child_text(node,"Status","ActivityStatus"),
                "task_type":_child_text(node,"Type","ActivityType") or "TT_Task",
                "target_start_date":_child_text(node,"PlannedStartDate","StartDate","Start"),
                "target_end_date":_child_text(node,"PlannedFinishDate","FinishDate","Finish"),
                "early_start_date":_child_text(node,"EarlyStartDate"),
                "early_end_date":_child_text(node,"EarlyFinishDate"),
                "act_start_date":_child_text(node,"ActualStartDate"),
                "act_end_date":_child_text(node,"ActualFinishDate"),
                "target_drtn_hr_cnt":_duration_hours(_child_text(node,"PlannedDuration","OriginalDuration")),
                "remain_drtn_hr_cnt":_duration_hours(_child_text(node,"RemainingDuration")),
                "total_float_hr_cnt":_duration_hours(_child_text(node,"TotalFloat")),
                "clndr_id":_child_text(node,"CalendarObjectId","CalendarId"),
                "cstr_type":_child_text(node,"PrimaryConstraintType","ConstraintType"),
                "cstr_date":_child_text(node,"PrimaryConstraintDate","ConstraintDate"),
            }
            tasks.append(task)
        elif tag=="relationship":
            pred=_child_text(node,"PredecessorActivityId","PredecessorActivityObjectId")
            succ=_child_text(node,"SuccessorActivityId","SuccessorActivityObjectId")
            if pred and succ:
                rels.append({"task_id":succ,"pred_task_id":pred,"pred_type":_child_text(node,"Type") or "PR_FS","lag_hr_cnt":_duration_hours(_child_text(node,"Lag")) or 0})
        elif tag=="wbs":
            wid=_child_text(node,"ObjectId","Id")
            if wid:wbs.append({"wbs_id":wid,"wbs_name":_child_text(node,"Name") or wid,"parent_wbs_id":_child_text(node,"ParentObjectId","ParentId")})
        elif tag=="project":
            projects.append({"proj_id":_child_text(node,"ObjectId","Id"),"proj_short_name":_child_text(node,"Id","ProjectId"),"proj_name":_child_text(node,"Name"),"plan_start_date":_child_text(node,"PlannedStartDate","StartDate"),"plan_end_date":_child_text(node,"ScheduledFinishDate","FinishDate"),"data_date":_child_text(node,"DataDate")})
        elif tag=="calendar":
            cid=_child_text(node,"ObjectId","Id")
            if cid:calendars.append({"clndr_id":cid,"clndr_name":_child_text(node,"Name") or cid})
        elif tag=="resource":
            rid=_child_text(node,"ObjectId","Id")
            if rid:resources.append({"rsrc_id":rid,"rsrc_name":_child_text(node,"Name"),"rsrc_type":_child_text(node,"ResourceType","Type")})
        elif tag in {"resourceassignment","resourceassignmentspread"}:
            tid=_child_text(node,"ActivityObjectId","ActivityId")
            rid=_child_text(node,"ResourceObjectId","ResourceId")
            if tid and rid:assignments.append({"task_id":tid,"rsrc_id":rid})
    if not tasks:return None
    return {"PROJECT":projects,"TASK":tasks,"TASKPRED":rels,"PROJWBS":wbs,"CALENDAR":calendars,"RSRC":resources,"TASKRSRC":assignments}


def _docx_schedule_tables(content:bytes):
    document=DocxDocument(io.BytesIO(content))
    candidates=[]
    for index,table in enumerate(document.tables,1):
        matrix=[[cell.text.strip() for cell in row.cells] for row in table.rows]
        parsed=_table_matrix_to_schedule(matrix,f"Word Table {index}")
        if parsed and parsed["TASK"]:candidates.append(parsed)
    if not candidates:return None
    tasks=[];wbs=[];rels=[]
    for parsed in candidates:
        tasks.extend(parsed["TASK"]);wbs.extend(parsed["PROJWBS"]);rels.extend(parsed["TASKPRED"])
    return {"PROJECT":[],"TASK":tasks,"PROJWBS":wbs,"TASKPRED":rels,"CALENDAR":[],"RSRC":[],"TASKRSRC":[]}


def _fixed_column_matrix(text:str):
    matrix=[]
    for raw in text.splitlines():
        line=raw.strip()
        if not line:continue
        cells=[x.strip() for x in re.split(r"\s{2,}",line) if x.strip()]
        if len(cells)>=2:matrix.append(cells)
    return matrix


def _report_text(extension:str,content:bytes):
    if extension=="docx":
        tabular=_docx_schedule_tables(content)
        if tabular:return tabular
    text=extract_text(extension,content)
    if not text.strip():return None
    matrices=[]
    delimited=[]
    for line in text.splitlines():
        if "\t" in line:delimited.append(line.split("\t"))
        elif "|" in line:delimited.append([x.strip() for x in line.split("|")])
    if delimited:matrices.append(delimited)
    fixed=_fixed_column_matrix(text)
    if fixed:matrices.append(fixed)
    for matrix in matrices:
        parsed=_table_matrix_to_schedule(matrix,"REPORT")
        if parsed and parsed["TASK"]:
            return {"PROJECT":[],"TASK":parsed["TASK"],"PROJWBS":parsed["PROJWBS"],"TASKPRED":parsed["TASKPRED"],"CALENDAR":[],"RSRC":[],"TASKRSRC":[]}
    return None


def ingest_schedule(extension:str,content:bytes):
    ext=extension.lower()
    if ext=="xer":
        tables=parse_xer(content);source_kind="Primavera P6 XER";fidelity="Full Native"
    elif ext=="xml":
        tables=_p6_xml(content);source_kind="Primavera / Schedule XML";fidelity="Structured"
    elif ext in {"xlsx","csv"}:
        tables=_spreadsheet(ext,content);source_kind="Schedule Spreadsheet";fidelity="Structured Table"
    elif ext in {"pdf","docx","txt"}:
        tables=_report_text(ext,content);source_kind="Schedule Report";fidelity="Report / Limited"
    elif ext=="mpp":
        tables=None;source_kind="Microsoft Project MPP";fidelity="Native Binary / Parser Required"
    elif ext in {"xls","doc"}:
        tables=None;source_kind=f"Legacy {ext.upper()} Schedule";fidelity="Legacy Binary / Conversion Required"
    else:
        tables=None;source_kind=f"{ext.upper()} Schedule";fidelity="Unsupported Structure"

    if not tables or not tables.get("TASK"):
        return {
            "detected":False,
            "tables":{},
            "source_kind":source_kind,
            "fidelity":fidelity,
            "capabilities":{"activities":False,"logic":False,"float":False,"resources":False,"calendars":False,"wbs":False},
            "limitations":["No reliable structured activity table could be extracted from this source."],
            "parser_version":"phase6-unified-schedule-ingestion-v2",
        }

    task_fields={k for row in tables.get("TASK",[]) for k,v in row.items() if v not in (None,"")}
    capabilities={
        "activities":bool(tables.get("TASK")),
        "logic":bool(tables.get("TASKPRED")),
        "float":"total_float_hr_cnt" in task_fields,
        "resources":bool(tables.get("TASKRSRC")),
        "calendars":bool(tables.get("CALENDAR")) or "clndr_id" in task_fields,
        "wbs":bool(tables.get("PROJWBS")) or "wbs_id" in task_fields,
    }
    limitations=[]
    for key,label in (("logic","network logic"),("float","total float"),("resources","resource assignments"),("calendars","calendar definitions"),("wbs","WBS structure")):
        if not capabilities[key]:limitations.append(f"{label.title()} is not available from the uploaded source.")
    return {
        "detected":True,
        "tables":tables,
        "source_kind":source_kind,
        "fidelity":fidelity,
        "capabilities":capabilities,
        "limitations":limitations,
        "parser_version":"phase6-unified-schedule-ingestion-v2",
    }
