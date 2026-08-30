import csv
import io
import re

from openpyxl import load_workbook


HEADER_SYNONYMS={
    "item_no":("item no","item number","item","sl no","sr no","serial no","boq item"),
    "description":("description","item description","description of work","scope","particulars","work description"),
    "unit":("unit","uom","unit of measurement"),
    "quantity":("quantity","qty","boq qty","estimated quantity"),
    "rate":("rate","unit rate","price"),
    "amount":("amount","total amount","value"),
    "work_front":("work front","workfront","location","section","zone","chainage","site","area"),
}


def _norm(value)->str:
    return re.sub(r"\s+"," ",str(value or "").strip().lower())


def _semantic(value)->str|None:
    text=_norm(value)
    for semantic,synonyms in HEADER_SYNONYMS.items():
        if text in synonyms:return semantic
    return None


def _detect_header(matrix:list[list]):
    best=None
    for row_index,row in enumerate(matrix[:50]):
        mapped={}
        for col_index,value in enumerate(row):
            semantic=_semantic(value)
            if semantic and semantic not in mapped:
                mapped[semantic]=col_index
        # Description is compulsory; require at least two additional structural fields.
        if "description" in mapped and len(mapped)>=3:
            score=len(mapped)+(1 if "item_no" in mapped else 0)+(1 if "quantity" in mapped else 0)
            if best is None or score>best[0]:
                best=(score,row_index,mapped)
    return best


def _rows_from_matrix(matrix:list[list],sheet_name:str):
    detected=_detect_header(matrix)
    if not detected:return [],None
    _,header_row,mapping=detected
    rows=[]
    blank_streak=0
    for values in matrix[header_row+1:]:
        description=str(values[mapping["description"]] if mapping["description"]<len(values) else "").strip()
        if not description:
            blank_streak+=1
            if blank_streak>=5:break
            continue
        blank_streak=0
        item_no=str(values[mapping["item_no"]] if "item_no" in mapping and mapping["item_no"]<len(values) else len(rows)+1).strip()
        unit=str(values[mapping["unit"]] if "unit" in mapping and mapping["unit"]<len(values) else "").strip() or None
        quantity=values[mapping["quantity"]] if "quantity" in mapping and mapping["quantity"]<len(values) else None
        rate=values[mapping["rate"]] if "rate" in mapping and mapping["rate"]<len(values) else None
        amount=values[mapping["amount"]] if "amount" in mapping and mapping["amount"]<len(values) else None
        work_front=str(values[mapping["work_front"]] if "work_front" in mapping and mapping["work_front"]<len(values) else "").strip() or None
        # Avoid obvious subtotal/header rows without quantity/unit when possible.
        lower=description.lower()
        if lower in {"total","sub total","subtotal","grand total"}:continue
        rows.append({
            "item_no":f"{sheet_name}:{item_no}" if sheet_name else item_no,
            "description":description,
            "unit":unit,
            "quantity":quantity,
            "rate":rate,
            "amount":amount,
            "work_front":work_front,
            "source_sheet":sheet_name,
        })
    return rows,{"sheet":sheet_name,"header_row":header_row+1,"columns":mapping}


def extract_boq_rows(extension:str,content:bytes):
    extension=extension.lower()
    all_rows=[];tables=[]
    if extension=="xlsx":
        workbook=load_workbook(io.BytesIO(content),data_only=False,read_only=True)
        for ws in workbook.worksheets:
            matrix=[list(row) for row in ws.iter_rows(values_only=True)]
            rows,table=_rows_from_matrix(matrix,ws.title)
            if rows:
                all_rows.extend(rows);tables.append({**table,"rows":len(rows)})
    elif extension=="csv":
        text=content.decode("utf-8-sig",errors="replace")
        matrix=list(csv.reader(io.StringIO(text)))
        rows,table=_rows_from_matrix(matrix,"")
        if rows:
            all_rows.extend(rows);tables.append({**table,"rows":len(rows)})
    else:
        return {"detected":False,"rows":[],"tables":[],"version":"phase6-boq-document-extractor-v1"}

    return {
        "detected":bool(all_rows),
        "rows":all_rows,
        "tables":tables,
        "summary":{
            "rows":len(all_rows),
            "tables":len(tables),
            "sheets":[x["sheet"] for x in tables if x.get("sheet")],
        },
        "version":"phase6-boq-document-extractor-v1",
    }
