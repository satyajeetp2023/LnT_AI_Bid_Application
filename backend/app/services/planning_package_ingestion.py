import csv
import io
import re
from datetime import date,datetime
from decimal import Decimal,InvalidOperation

import xlrd
from openpyxl import load_workbook
from sqlalchemy import delete,select
from sqlalchemy.orm import Session

from app.models import AuditEvent,BidDocument,PlanningResourceEntry


PLAN_EXTENSIONS={"xlsx","xls","csv"}

ALIASES={
 "resource_name":("resource","resource name","equipment","equipment name","plant","p&m","machinery","staff","staff name","designation","position","role","trade","manpower"),
 "role_or_trade":("designation","position","role","trade","category","skill","staff category","manpower category"),
 "activity_reference":("activity","activity name","activity id","task","task name","work item","boq item","scope"),
 "work_front":("work front","front","section","location","zone","area","chainage"),
 "quantity":("qty","quantity","nos","no.","number","strength","headcount","count","units"),
 "unit":("unit","uom"),
 "start_date":("start","start date","from","mobilization","mobilisation","mobilization date","mobilisation date","deployment date"),
 "finish_date":("finish","finish date","to","end","end date","demobilization","demobilisation","release date"),
 "shift_hours":("shift hours","hours per shift","working hours","hrs/day","hours/day"),
 "productivity_rate":("productivity","productivity rate","output","production rate","rate/day","qty/day"),
 "productivity_unit":("productivity unit","output unit","rate unit"),
 "notes":("notes","remarks","remark","comments","comment"),
}

STAFF_SIGNALS=("staff","engineer","manager","supervisor","planner","planning","qa","qc","quality","safety","hse","commercial","contracts","surveyor","draught","draftsman","coordinator")
EQUIPMENT_SIGNALS=("equipment","plant","p&m","machinery","crane","excavator","loader","vehicle","truck","tower wagon","tamping","machine","generator","compressor")
LABOUR_SIGNALS=("labour","labor","manpower","crew","gang","technician","fitter","electrician","welder","helper","skilled","unskilled")


def _norm(value):
 return re.sub(r"\s+"," ",str(value or "").strip().lower())


def _semantic(value):
 text=_norm(value).strip(" :*-")
 for key,names in ALIASES.items():
  if text in names:return key
  if any(text.startswith(name+" ") for name in names if len(name)>=5):return key
 return None


def _decimal(value):
 if value in (None,""):return None
 if isinstance(value,(int,float,Decimal)):return Decimal(str(value))
 text=re.sub(r"[^0-9.\-]","",str(value))
 if not text:return None
 try:return Decimal(text)
 except InvalidOperation:return None


def _date(value):
 if value in (None,""):return None
 if isinstance(value,datetime):return value.date()
 if isinstance(value,date):return value
 if isinstance(value,(int,float)):
  try:return xlrd.xldate_as_datetime(value,0).date()
  except Exception:return None
 text=str(value).strip()
 for fmt in ("%d-%m-%Y","%d/%m/%Y","%Y-%m-%d","%d-%b-%Y","%d %b %Y","%b-%Y","%b %Y"):
  try:return datetime.strptime(text,fmt).date()
  except ValueError:pass
 return None


def _plan_type(filename:str,sheet:str,headers:list[str],rows:list[list]):
 hay=_norm(" ".join([filename,sheet,*headers]))
 sample=_norm(" ".join(str(x) for row in rows[:20] for x in row if x not in (None,"")))
 text=f"{hay} {sample}"
 if any(x in text for x in ("staff plan","staff deployment","staffing plan","key personnel","organization chart","organisation chart")):
  return "Staff Plan"
 if any(x in text for x in ("equipment plan","plant plan","p&m plan","machinery plan","equipment deployment")):
  return "Equipment Plan"
 if any(x in text for x in ("resource plan","resource deployment","manpower plan","labour plan","labor plan","crew plan")):
  return "Resource Plan"
 # Header/content inference
 if any(x in text for x in ("designation","position","staff category")) and not any(x in text for x in EQUIPMENT_SIGNALS):
  return "Staff Plan"
 if any(x in text for x in EQUIPMENT_SIGNALS):
  return "Equipment Plan"
 return "Resource Plan"


def _resource_category(plan_type:str,name:str,role:str|None):
 text=_norm(f"{name} {role or ''}")
 if plan_type=="Staff Plan" or any(x in text for x in STAFF_SIGNALS):return "Staff"
 if plan_type=="Equipment Plan" or any(x in text for x in EQUIPMENT_SIGNALS):return "Equipment"
 if any(x in text for x in LABOUR_SIGNALS):return "Labour"
 return "Other"


def _sheet_matrices(extension:str,content:bytes):
 if extension=="xlsx":
  wb=load_workbook(io.BytesIO(content),data_only=True,read_only=True)
  for ws in wb.worksheets:
   yield ws.title,[list(x) for x in ws.iter_rows(values_only=True)]
 elif extension=="xls":
  wb=xlrd.open_workbook(file_contents=content)
  for ws in wb.sheets():
   yield ws.name,[ws.row_values(r) for r in range(ws.nrows)]
 elif extension=="csv":
  rows=list(csv.reader(io.StringIO(content.decode("utf-8-sig",errors="replace"))))
  yield "CSV",rows


def _find_header(matrix:list[list]):
 best=None
 for row_index,row in enumerate(matrix[:60]):
  mapping={}
  for col_index,value in enumerate(row):
   semantic=_semantic(value)
   if semantic and semantic not in mapping:mapping[semantic]=col_index
  # At least a resource/staff identity plus one planning dimension.
  if "resource_name" in mapping and any(x in mapping for x in ("quantity","start_date","finish_date","activity_reference","work_front")):
   score=len(mapping)
   if best is None or score>best[0]:best=(score,row_index,mapping)
 return best


def detect_planning_resource_document(filename:str,extension:str,content:bytes):
 ext=extension.lower()
 if ext not in PLAN_EXTENSIONS:return {"detected":False,"reason":"Unsupported format"}
 candidates=[]
 for sheet,matrix in _sheet_matrices(ext,content):
  header=_find_header(matrix)
  if not header:continue
  _,row_index,mapping=header
  headers=[str(x or "") for x in matrix[row_index]]
  candidates.append({
   "sheet":sheet,"header_row":row_index+1,"mapping":mapping,
   "plan_type":_plan_type(filename,sheet,headers,matrix[row_index+1:]),
  })
 if not candidates:return {"detected":False,"reason":"No resource/staff planning table detected","sheets":[]}
 kinds=sorted({x["plan_type"] for x in candidates})
 return {"detected":True,"plan_types":kinds,"sheets":candidates,"parser_version":"planning-package-ingestion-v1"}


def ingest_planning_resource_document(
 db:Session,document:BidDocument,storage,user_id:int,request_metadata:dict|None=None
):
 if not document.storage_path:raise ValueError("Document content is not available")
 ext=document.file_extension.lower()
 if ext not in PLAN_EXTENSIONS:return {"detected":False,"created":0,"reason":"Unsupported format"}
 content=storage.read(document.storage_path)
 detection=detect_planning_resource_document(document.original_filename,ext,content)
 if not detection["detected"]:return {**detection,"created":0}

 db.execute(delete(PlanningResourceEntry).where(PlanningResourceEntry.source_document_id==document.id))
 created=0;plan_types=set()
 for sheet,matrix in _sheet_matrices(ext,content):
  header=_find_header(matrix)
  if not header:continue
  _,header_idx,mapping=header
  headers=[str(x or "") for x in matrix[header_idx]]
  plan_type=_plan_type(document.original_filename,sheet,headers,matrix[header_idx+1:])
  plan_types.add(plan_type)
  for offset,row in enumerate(matrix[header_idx+1:],start=header_idx+2):
   def get(name):
    col=mapping.get(name)
    return row[col] if col is not None and col<len(row) else None
   name=str(get("resource_name") or "").strip()
   if not name:continue
   role=str(get("role_or_trade") or "").strip() or None
   quantity=_decimal(get("quantity"))
   start=_date(get("start_date"));finish=_date(get("finish_date"))
   # Ignore obvious total/footer rows with no planning context.
   if _norm(name) in {"total","grand total","subtotal"}:continue
   entry=PlanningResourceEntry(
    bid_project_id=document.bid_project_id,source_document_id=document.id,
    source_sheet=sheet,source_row=offset,plan_type=plan_type,
    resource_category=_resource_category(plan_type,name,role),resource_name=name,
    role_or_trade=role,
    activity_reference=str(get("activity_reference") or "").strip() or None,
    work_front=str(get("work_front") or "").strip() or None,
    quantity=quantity,unit=str(get("unit") or "").strip() or None,
    start_date=start,finish_date=finish,
    shift_hours=_decimal(get("shift_hours")),
    productivity_rate=_decimal(get("productivity_rate")),
    productivity_unit=str(get("productivity_unit") or "").strip() or None,
    notes=str(get("notes") or "").strip() or None,
    extraction_confidence=Decimal(".90" if len(mapping)>=5 else ".78"),
    created_by=user_id,
   )
   db.add(entry);created+=1

 if created:
  document.document_category="Forms / Formats / Schedules"
  if len(plan_types)==1:document.document_type=next(iter(plan_types))
  else:document.document_type="Integrated Resource / Staff Plan"
  document.classification_status="classified"
  document.document_status="Uploaded"
 db.add(AuditEvent(
  user_id=user_id,bid_project_id=document.bid_project_id,event_type="planning_package.resource_plan_ingested",
  entity_type="BidDocument",entity_id=str(document.id),request_metadata=request_metadata or {},
  details={"created":created,"plan_types":sorted(plan_types),"parser_version":"planning-package-ingestion-v1"},
 ))
 db.commit()
 return {
  "detected":True,"created":created,"plan_types":sorted(plan_types),
  "document_type":document.document_type,"parser_version":"planning-package-ingestion-v1",
 }


def planning_resource_summary(db:Session,bid_id:int):
 rows=db.scalars(select(PlanningResourceEntry).where(
  PlanningResourceEntry.bid_project_id==bid_id
 ).order_by(PlanningResourceEntry.plan_type,PlanningResourceEntry.resource_name)).all()
 return {
  "items":[{
   "id":x.id,"source_document_id":x.source_document_id,"source_sheet":x.source_sheet,"source_row":x.source_row,
   "plan_type":x.plan_type,"resource_category":x.resource_category,"resource_name":x.resource_name,
   "role_or_trade":x.role_or_trade,"activity_reference":x.activity_reference,"work_front":x.work_front,
   "quantity":float(x.quantity) if x.quantity is not None else None,"unit":x.unit,
   "start_date":x.start_date.isoformat() if x.start_date else None,"finish_date":x.finish_date.isoformat() if x.finish_date else None,
   "shift_hours":float(x.shift_hours) if x.shift_hours is not None else None,
   "productivity_rate":float(x.productivity_rate) if x.productivity_rate is not None else None,
   "productivity_unit":x.productivity_unit,"notes":x.notes,"confidence":float(x.extraction_confidence),
  } for x in rows],
  "summary":{
   "entries":len(rows),
   "staff_entries":sum(1 for x in rows if x.resource_category=="Staff"),
   "labour_entries":sum(1 for x in rows if x.resource_category=="Labour"),
   "equipment_entries":sum(1 for x in rows if x.resource_category=="Equipment"),
   "dated_entries":sum(1 for x in rows if x.start_date or x.finish_date),
   "activity_linked_entries":sum(1 for x in rows if x.activity_reference),
   "work_front_entries":sum(1 for x in rows if x.work_front),
   "plan_types":sorted({x.plan_type for x in rows}),
  },
  "version":"planning-package-ingestion-v1",
 }
