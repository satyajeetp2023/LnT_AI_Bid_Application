import csv
import io
import re
from decimal import Decimal,InvalidOperation

from openpyxl import load_workbook
from pypdf import PdfReader


ALIASES={
 "rank":{"rank","position","l rank","l1/l2/l3/l4","bid rank"},
 "bidder_name":{"bidder","bidder name","contractor","participant","company","tenderer"},
 "bid_value":{"bid value","quoted value","price","quoted price","tender value","amount"},
 "currency":{"currency","curr"},
 "is_ours":{"our bid","is ours","ours","l&t","our offer"},
}


def _norm(value):
 return re.sub(r"\s+"," ",str(value or "").strip().lower())


def _semantic(value):
 text=_norm(value)
 for key,names in ALIASES.items():
  if text in names:return key
 return None


def _rank(value):
 if value in (None,""):return None
 text=str(value).strip().upper()
 match=re.search(r"(?:L\s*)?(\d{1,3})",text)
 return int(match.group(1)) if match else None


def _money(value):
 if value in (None,""):return None
 text=re.sub(r"[^0-9.()-]","",str(value)).replace("(","-").replace(")","")
 try:return Decimal(text)
 except (InvalidOperation,ValueError):return None


def _truthy(value):
 return _norm(value) in {"yes","y","true","1","ours","our bid","l&t","lt","l and t"}

def _finalize(prices,warnings,filename,parser_version):
 prices.sort(key=lambda x:x["rank"])
 ours=[x for x in prices if x["is_ours"]]
 if len(ours)>1:warnings.append("More than one row appears to be marked as our bid.")
 outcome={
  "result_status":"Won" if len(ours)==1 and ours[0]["rank"]==1 else "Lost" if len(ours)==1 else "Result Awaited",
  "our_rank":ours[0]["rank"] if len(ours)==1 else None,
  "our_bid_value":ours[0]["bid_value"] if len(ours)==1 else None,
  "awarded_bidder":next((x["bidder_name"] for x in prices if x["rank"]==1),None),
  "source_reference":filename or None,
 }
 return {
  "detected":bool(prices),"source_filename":filename,"outcome_candidate":outcome,"prices":prices,
  "warnings":warnings,"requires_review":True,
  "note":"Preview only. No bid outcome or bidder price is saved until an authorized user reviews and explicitly saves it.",
  "parser_version":parser_version,
 }


def _preview_pdf_text(text:str,filename:str):
 if len(text.strip())<40:
  return {
   "detected":False,"source_filename":filename,"prices":[],
   "warnings":["PDF has insufficient extractable text. Scanned/image result notices require vision/OCR review."],
   "requires_review":True,"parser_version":"phase7-historical-result-pdf-v1",
  }
 prices=[];warnings=[];seen_ranks=set();seen_bidders=set()
 pattern=re.compile(r"^\s*(?:L\s*)?(\d{1,3})\s+(.+?)\s+([0-9][0-9,]*(?:\.[0-9]+)?)\s*(INR|USD|EUR|GBP)?\s*$",re.I)
 for line_no,line in enumerate(text.splitlines(),start=1):
  match=pattern.match(line.strip())
  if not match:continue
  rank=int(match.group(1));bidder=match.group(2).strip(" -|");value=_money(match.group(3));currency=(match.group(4) or "INR").upper()
  if rank>100 or not bidder or value is None:continue
  if rank in seen_ranks:
   warnings.append(f"Duplicate rank {rank} detected at PDF text line {line_no}; review required.");continue
  key=bidder.lower()
  if key in seen_bidders:
   warnings.append(f"Duplicate bidder '{bidder}' detected at PDF text line {line_no}; review required.");continue
  seen_ranks.add(rank);seen_bidders.add(key)
  prices.append({"bidder_name":bidder,"rank":rank,"bid_value":float(value),"currency":currency,"is_ours":_norm(bidder) in {"l&t","larsen & toubro","larsen and toubro","lt"},"source_reference":filename or None})
 if not prices:
  warnings.append("No ranked bidder-price rows could be read from the PDF text; manual review is required.")
 return _finalize(prices,warnings,filename,"phase7-historical-result-pdf-v1")


def _preview_pdf(content:bytes,filename:str):
 try:
  reader=PdfReader(io.BytesIO(content))
  text="\n".join(page.extract_text() or "" for page in reader.pages)
 except Exception:
  return {"detected":False,"source_filename":filename,"prices":[],"warnings":["PDF could not be safely parsed; manual review is required."],"requires_review":True,"parser_version":"phase7-historical-result-pdf-v1"}
 return _preview_pdf_text(text,filename)



def _matrix(extension,content):
 if extension=="csv":
  return list(csv.reader(io.StringIO(content.decode("utf-8-sig",errors="replace"))))
 if extension=="xlsx":
  wb=load_workbook(io.BytesIO(content),data_only=True,read_only=True)
  ws=wb.active
  return [list(row) for row in ws.iter_rows(values_only=True)]
 raise ValueError("Only CSV and XLSX tender result tables are supported for automatic preview.")


def preview_historical_result(extension:str,content:bytes,filename:str=""):
 extension=extension.lower()
 if extension=="pdf":return _preview_pdf(content,filename)
 matrix=_matrix(extension,content)
 best=None
 for row_index,row in enumerate(matrix[:50]):
  mapping={}
  for col_index,value in enumerate(row):
   semantic=_semantic(value)
   if semantic and semantic not in mapping:mapping[semantic]=col_index
  if {"rank","bidder_name","bid_value"}<=set(mapping):
   score=len(mapping)
   if best is None or score>best[0]:best=(score,row_index,mapping)
 if not best:
  return {
   "detected":False,"source_filename":filename,"prices":[],"warnings":["Could not identify Rank, Bidder and Bid Value columns."],
   "requires_review":True,"parser_version":"phase7-historical-result-preview-v1",
  }
 _,header_row,mapping=best
 prices=[];warnings=[]
 seen_ranks=set();seen_bidders=set()
 for row_no,row in enumerate(matrix[header_row+1:],start=header_row+2):
  def get(name):
   col=mapping.get(name);return row[col] if col is not None and col<len(row) else None
  rank=_rank(get("rank"));bidder=str(get("bidder_name") or "").strip();value=_money(get("bid_value"))
  if rank is None and not bidder and value is None:continue
  if rank is None or not bidder or value is None:
   warnings.append(f"Row {row_no} was skipped because rank, bidder or bid value is incomplete.")
   continue
  if rank in seen_ranks:
   warnings.append(f"Duplicate rank {rank} detected at row {row_no}; review required.")
   continue
  key=bidder.lower()
  if key in seen_bidders:
   warnings.append(f"Duplicate bidder '{bidder}' detected at row {row_no}; review required.")
   continue
  seen_ranks.add(rank);seen_bidders.add(key)
  currency=str(get("currency") or "INR").strip().upper()[:3] or "INR"
  is_ours=_truthy(get("is_ours")) or _norm(bidder) in {"l&t","larsen & toubro","larsen and toubro","lt"}
  prices.append({
   "bidder_name":bidder,"rank":rank,"bid_value":float(value),"currency":currency,
   "is_ours":is_ours,"source_reference":filename or None,
  })
 return _finalize(prices,warnings,filename,"phase7-historical-result-preview-v1")
