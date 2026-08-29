import io
import re
from dataclasses import asdict,dataclass

from openpyxl import load_workbook


YES_NO_SIGNALS={"yes","no","compliant","not compliant","compliance"}
FIELD_HINTS=("tenderer","document","name","date","reference","clause","comment","proposal","signature","stamp","designation")
TABLE_HINTS=("clause","yes","no","comments","proposal","description","item","response","compliance")


@dataclass
class TemplateCell:
    coordinate:str
    row:int
    column:int
    value:str|None
    merged_range:str|None
    role:str
    confidence:float


@dataclass
class TemplateTable:
    sheet:str
    header_rows:list[int]
    start_row:int
    end_row:int
    columns:list[dict]
    table_type:str
    confidence:float


def _norm(value)->str:
    return re.sub(r"\s+"," ",str(value or "")).strip()


def _merged_range(ws,coordinate:str)->str|None:
    for rng in ws.merged_cells.ranges:
        if coordinate in rng:return str(rng)
    return None


def _cell_role(value:str,has_border:bool,merged:bool)->tuple[str,float]:
    lower=value.lower()
    if any(x in lower for x in FIELD_HINTS) and (":" in value or len(value)<80):
        return "label",.86
    if lower in YES_NO_SIGNALS:
        return "choice_header",.96
    if has_border and not value:
        return "candidate_input",.74
    if merged and len(value)>6:
        return "heading",.78
    if value:
        return "text",.60
    return "blank",.20


def _table_type(headers:list[str])->tuple[str,float]:
    joined=" ".join(headers).lower()
    if "clause" in joined and "yes" in joined and "no" in joined:
        return "statement_of_compliance",.98
    if "clause" in joined and ("comment" in joined or "proposal" in joined):
        return "compliance_or_deviation_register",.92
    hits=sum(1 for x in TABLE_HINTS if x in joined)
    if hits>=3:return "structured_submission_table",min(.90,.55+hits*.07)
    return "generic_table",.55


def _semantic_field(header:str)->tuple[str,str]:
    lower=header.lower()
    if "clause" in lower and "reference" in lower:return "clause_reference","reference"
    if "yes" in lower and ("compliant" in lower or "compliance" in lower):return "compliant_yes","bidder_input"
    if "no" in lower and ("compliant" in lower or "compliance" in lower):return "compliant_no","bidder_input"
    if ("comment" in lower or "proposal" in lower) and "tender" in lower:return "tenderer_comments","bidder_input"
    if "evaluator" in lower and ("remark" in lower or "comment" in lower):return "evaluator_remarks","employer_only"
    if "remark" in lower or "comment" in lower:return "comments","bidder_input"
    return re.sub(r"[^a-z0-9]+","_",lower).strip("_") or "field","unknown"


def _header_block(ws,row_index:int,max_depth:int=2):
    values={}
    header_rows=[]
    for depth in range(max_depth):
        current=row_index+depth
        if current>ws.max_row:break
        row_has_signal=False
        for col_index in range(1,ws.max_column+1):
            value=_norm(ws.cell(current,col_index).value)
            if value:
                values.setdefault(col_index,[]).append(value)
                if any(x in value.lower() for x in TABLE_HINTS) or value.lower() in YES_NO_SIGNALS:
                    row_has_signal=True
            else:
                for rng in ws.merged_cells.ranges:
                    if current>=rng.min_row and current<=rng.max_row and col_index>=rng.min_col and col_index<=rng.max_col:
                        anchor=_norm(ws.cell(rng.min_row,rng.min_col).value)
                        if anchor and anchor not in values.setdefault(col_index,[]):values[col_index].append(anchor)
                        break
        if depth==0 or row_has_signal:
            header_rows.append(current)
        elif depth>0:
            break
    headers=[]
    columns=[]
    for col_index,parts in sorted(values.items()):
        joined=" ".join(dict.fromkeys(parts))
        if not joined:continue
        semantic,ownership=_semantic_field(joined)
        headers.append(joined)
        columns.append({
            "column":col_index,
            "coordinate":ws.cell(row_index,col_index).coordinate,
            "header":joined,
            "semantic_field":semantic,
            "ownership":ownership,
        })
    return header_rows,headers,columns


def _structure_signature(table:dict)->tuple:
    return (
        table["table_type"],
        tuple((c["column"],c.get("semantic_field"),c.get("ownership")) for c in table["columns"]),
        len(table["header_rows"]),
    )


PLACEHOLDER_RE=re.compile(r"\b([A-Z][A-Z0-9 /&().]{2,40})\s*[-_]{5,}")


def _placeholder_semantic(label:str)->tuple[str,str]:
    lower=label.lower().strip()
    if "tenderer" in lower or "bidder" in lower:return "tenderer_name","bidder_master_data"
    if "document" in lower:return "document_reference","bid_specific_input"
    if "date" in lower:return "date","bid_specific_input"
    if "signature" in lower:return "signature","manual_controlled_input"
    return re.sub(r"[^a-z0-9]+","_",lower).strip("_") or "placeholder","bid_specific_input"


def _inline_placeholders(value:str,coordinate:str):
    result=[]
    for match in PLACEHOLDER_RE.finditer(value):
        label=match.group(1).strip()
        semantic,source=_placeholder_semantic(label)
        result.append({
            "label":label,
            "semantic_field":semantic,
            "input_source":source,
            "coordinate":coordinate,
        })
    return result


def parse_xlsx_template(content:bytes)->dict:
    workbook=load_workbook(io.BytesIO(content),data_only=False)
    sheets=[]
    all_tables=[]
    for ws in workbook.worksheets:
        cells=[]
        nonempty=[]
        placeholders=[]
        for row in ws.iter_rows():
            for cell in row:
                value=_norm(cell.value)
                merged_range=_merged_range(ws,cell.coordinate)
                has_border=any(getattr(side,"style",None) for side in (cell.border.left,cell.border.right,cell.border.top,cell.border.bottom))
                role,confidence=_cell_role(value,has_border,bool(merged_range))
                if value or has_border:
                    cells.append(asdict(TemplateCell(cell.coordinate,cell.row,cell.column,value or None,merged_range,role,confidence)))
                if value:
                    nonempty.append(cell)
                    placeholders.extend(_inline_placeholders(value,cell.coordinate))

        tables=[]
        used_header_rows=set()
        for row_index in range(1,ws.max_row+1):
            if row_index in used_header_rows:continue
            header_rows,headers,columns=_header_block(ws,row_index,2)
            if len(headers)<2:continue
            lower=" ".join(headers).lower()
            hits=sum(1 for x in TABLE_HINTS if x in lower)
            if hits<2:continue
            table_type,confidence=_table_type(headers)
            if table_type=="statement_of_compliance" and len(header_rows)>1:confidence=.99
            start_row=max(header_rows)+1
            end_row=max(header_rows)
            for scan in range(start_row,min(ws.max_row,start_row+50)+1):
                if any(_norm(ws.cell(scan,c["column"]).value) for c in columns) or any(
                    any(getattr(side,"style",None) for side in (ws.cell(scan,c["column"]).border.left,ws.cell(scan,c["column"]).border.right,ws.cell(scan,c["column"]).border.top,ws.cell(scan,c["column"]).border.bottom))
                    for c in columns
                ):
                    end_row=scan
                elif end_row>=start_row:
                    break
            table=asdict(TemplateTable(ws.title,header_rows,start_row,end_row,columns,table_type,confidence))
            tables.append(table);all_tables.append(table)
            used_header_rows.update(header_rows)

        sheets.append({
            "name":ws.title,
            "max_row":ws.max_row,
            "max_column":ws.max_column,
            "merged_ranges":[str(x) for x in ws.merged_cells.ranges],
            "cells":cells,
            "tables":tables,
            "placeholders":placeholders,
            "image_count":len(ws._images),
        })

    signatures={}
    for table in all_tables:
        signatures.setdefault(_structure_signature(table),[]).append(table)
    workbook_patterns=[]
    for signature,tables in signatures.items():
        if len(tables)<3:continue
        clause_refs=[]
        for table in tables:
            clause_col=next((c["column"] for c in table["columns"] if c.get("semantic_field")=="clause_reference"),None)
            if clause_col:
                value=_norm(workbook[table["sheet"]].cell(table["start_row"],clause_col).value)
                if value:clause_refs.append(value)
        if len(clause_refs)>=3:
            workbook_patterns.append({
                "pattern_type":"repeated_sheet_per_clause",
                "table_type":tables[0]["table_type"],
                "sheet_count":len(tables),
                "clause_reference_count":len(clause_refs),
                "sample_clause_references":clause_refs[:10],
                "confidence":.99 if len(tables)>=10 else .94,
            })

    placeholder_groups={}
    for sheet in sheets:
        for item in sheet.get("placeholders",[]):
            key=item["semantic_field"]
            group=placeholder_groups.setdefault(key,{
                "semantic_field":item["semantic_field"],
                "label":item["label"],
                "input_source":item["input_source"],
                "occurrences":[],
            })
            group["occurrences"].append({"sheet":sheet["name"],"coordinate":item["coordinate"]})
    workbook_placeholders=[]
    for group in placeholder_groups.values():
        group["occurrence_count"]=len(group["occurrences"])
        group["sample_occurrences"]=group["occurrences"][:10]
        del group["occurrences"]
        workbook_placeholders.append(group)
    workbook_placeholders.sort(key=lambda x:x["semantic_field"])

    return {
        "file_type":"xlsx",
        "sheet_count":len(workbook.worksheets),
        "sheets":sheets,
        "tables":all_tables,
        "workbook_patterns":workbook_patterns,
        "workbook_placeholders":workbook_placeholders,
        "summary":{
            "tables_detected":len(all_tables),
            "compliance_tables":sum(1 for x in all_tables if x["table_type"]=="statement_of_compliance"),
            "candidate_input_cells":sum(1 for s in sheets for c in s["cells"] if c["role"]=="candidate_input"),
            "images_detected":sum(s["image_count"] for s in sheets),
            "repeated_sheet_patterns":len(workbook_patterns),
            "workbook_placeholders":len(workbook_placeholders),
        },
        "parser_version":"phase5-xlsx-template-parser-v3",
    }
