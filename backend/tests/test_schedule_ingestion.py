import io

from docx import Document
from openpyxl import Workbook

from app.services.p6_xer import analyze_schedule_tables
from app.services.p6_schedule_optimizer import build_schedule_optimization_from_tables
from app.services.schedule_ingestion import ingest_schedule


def test_excel_schedule_normalizes_without_inventing_logic_or_float():
    wb=Workbook()
    ws=wb.active
    ws.title="Programme"
    ws.append(["Activity ID","Activity Name","WBS","Start","Finish","Duration"])
    ws.append(["A100","Foundation","Civil","2026-01-01","2026-01-10",10])
    ws.append(["A200","Mast Erection","OHE","2026-01-11","2026-01-20",10])
    stream=io.BytesIO();wb.save(stream)

    ing=ingest_schedule("xlsx",stream.getvalue())
    assert ing["detected"] is True
    assert ing["capabilities"]["activities"] is True
    assert ing["capabilities"]["logic"] is False
    assert ing["capabilities"]["float"] is False

    analysis=analyze_schedule_tables(ing["tables"],capabilities=ing["capabilities"])
    assert analysis["health"]["issue_counts"]["open_start"]==0
    assert analysis["health"]["issue_counts"]["open_finish"]==0
    assert analysis["health"]["issue_counts"]["negative_float"]==0
    assert analysis["criticality"]["critical_activities"]==[]

    advisor=build_schedule_optimization_from_tables(ing["tables"],capabilities=ing["capabilities"])
    assert all("open start" not in " ".join(x["issues"]).lower() for x in advisor["optimization"]["candidates"])
    assert all("negative total float" not in x["issues"] for x in advisor["optimization"]["candidates"])


def test_primavera_style_xml_normalizes_activity_and_relationship():
    xml=b"""<?xml version="1.0"?>
<Project>
  <Name>Railway Test</Name>
  <Activity>
    <ObjectId>1</ObjectId><Id>A100</Id><Name>Foundation</Name>
    <PlannedStartDate>2026-01-01</PlannedStartDate>
    <PlannedFinishDate>2026-01-10</PlannedFinishDate>
    <OriginalDuration>PT80H</OriginalDuration><TotalFloat>PT40H</TotalFloat>
  </Activity>
  <Activity>
    <ObjectId>2</ObjectId><Id>A200</Id><Name>Mast Erection</Name>
    <PlannedStartDate>2026-01-11</PlannedStartDate>
    <PlannedFinishDate>2026-01-20</PlannedFinishDate>
    <OriginalDuration>PT80H</OriginalDuration><TotalFloat>PT20H</TotalFloat>
  </Activity>
  <Relationship>
    <PredecessorActivityId>A100</PredecessorActivityId>
    <SuccessorActivityId>A200</SuccessorActivityId>
    <Type>Finish to Start</Type>
  </Relationship>
</Project>"""
    ing=ingest_schedule("xml",xml)
    assert ing["detected"] is True
    assert len(ing["tables"]["TASK"])==2
    assert len(ing["tables"]["TASKPRED"])==1
    assert ing["capabilities"]["logic"] is True
    assert ing["capabilities"]["float"] is True


def test_iso_duration_is_parsed_as_hours_not_days():
    xml=b"""<?xml version="1.0"?>
<Project>
  <Activity>
    <Id>A100</Id><Name>Foundation</Name>
    <OriginalDuration>PT80H</OriginalDuration>
    <TotalFloat>PT16H</TotalFloat>
  </Activity>
</Project>"""
    ing=ingest_schedule("xml",xml)
    task=ing["tables"]["TASK"][0]
    assert task["target_drtn_hr_cnt"]==80
    assert task["total_float_hr_cnt"]==16


def test_word_schedule_table_is_extracted():
    document=Document()
    table=document.add_table(rows=1,cols=6)
    headers=["Activity ID","Activity Name","WBS","Start Date","Finish Date","Duration"]
    for i,value in enumerate(headers):
        table.rows[0].cells[i].text=value
    row=table.add_row().cells
    values=["A100","OHE Foundation","OHE","2026-01-01","2026-01-10","10 days"]
    for i,value in enumerate(values):
        row[i].text=value
    stream=io.BytesIO();document.save(stream)
    ing=ingest_schedule("docx",stream.getvalue())
    assert ing["detected"] is True
    assert ing["fidelity"]=="Report / Limited"
    assert ing["tables"]["TASK"][0]["task_code"]=="A100"
    assert ing["tables"]["TASK"][0]["target_drtn_hr_cnt"]==80


def test_mpp_is_recognized_as_native_binary_but_not_falsely_parsed():
    ing=ingest_schedule("mpp",b"not-a-real-mpp")
    assert ing["detected"] is False
    assert ing["source_kind"]=="Microsoft Project MPP"
    assert ing["fidelity"]=="Native Binary / Parser Required"


def test_excel_predecessor_lag_resources_and_milestone_are_normalized():
    wb=Workbook()
    ws=wb.active
    ws.title="Programme"
    ws.append(["Activity ID","Activity Name","Duration","Predecessors","Resources"])
    ws.append(["A100","Foundation","10 days","","Civil Crew"])
    ws.append(["A200","Mast Erection","5 days","A100FS+2d","OHE Crew; Crane"])
    ws.append(["M100","Section Complete","0 days","A200FF-4h",""])
    stream=io.BytesIO();wb.save(stream)

    ing=ingest_schedule("xlsx",stream.getvalue())
    assert ing["detected"] is True
    tasks={x["task_code"]:x for x in ing["tables"]["TASK"]}
    assert tasks["M100"]["task_type"]=="TT_Mile"

    rels={(x["task_id"],x["pred_task_id"]):x for x in ing["tables"]["TASKPRED"]}
    assert rels[("A200","A100")]["pred_type"]=="PR_FS"
    assert rels[("A200","A100")]["lag_hr_cnt"]==16
    assert rels[("M100","A200")]["pred_type"]=="PR_FF"
    assert rels[("M100","A200")]["lag_hr_cnt"]==-4

    names={x["rsrc_name"] for x in ing["tables"]["RSRC"]}
    assert {"Civil Crew","OHE Crew","Crane"}<=names
    assert len(ing["tables"]["TASKRSRC"])==3
    assert ing["capabilities"]["resources"] is True


def test_microsoft_project_xml_semantics_are_normalized():
    xml=b"""<?xml version="1.0"?>
<Project xmlns="http://schemas.microsoft.com/project">
  <Name>MSP Railway Test</Name>
  <Tasks>
    <Task>
      <UID>10</UID><ID>1</ID><Name>Foundation</Name><WBS>1.1</WBS>
      <Start>2026-01-01T08:00:00</Start><Finish>2026-01-10T17:00:00</Finish>
      <Duration>PT80H</Duration><TotalSlack>PT24H</TotalSlack>
      <PercentComplete>100</PercentComplete><CalendarUID>1</CalendarUID>
    </Task>
    <Task>
      <UID>20</UID><ID>2</ID><Name>Section Complete</Name><WBS>1.2</WBS>
      <Start>2026-01-11T08:00:00</Start><Finish>2026-01-11T08:00:00</Finish>
      <Duration>PT0H</Duration><Milestone>1</Milestone><PercentComplete>0</PercentComplete>
      <PredecessorLink>
        <PredecessorUID>10</PredecessorUID><Type>1</Type><LinkLag>600</LinkLag>
      </PredecessorLink>
    </Task>
  </Tasks>
  <Calendars><Calendar><UID>1</UID><Name>Standard</Name></Calendar></Calendars>
  <Resources><Resource><UID>5</UID><Name>OHE Crew</Name><Type>1</Type></Resource></Resources>
  <Assignments><Assignment><TaskUID>20</TaskUID><ResourceUID>5</ResourceUID></Assignment></Assignments>
</Project>"""
    ing=ingest_schedule("xml",xml)
    assert ing["detected"] is True
    assert ing["source_kind"]=="Microsoft Project XML"
    tasks={x["task_id"]:x for x in ing["tables"]["TASK"]}
    assert tasks["10"]["target_drtn_hr_cnt"]==80
    assert tasks["10"]["total_float_hr_cnt"]==24
    assert tasks["10"]["status_code"]=="Complete"
    assert tasks["20"]["task_type"]=="TT_Mile"
    rel=ing["tables"]["TASKPRED"][0]
    assert rel["task_id"]=="20"
    assert rel["pred_task_id"]=="10"
    assert rel["pred_type"]=="PR_FS"
    assert rel["lag_hr_cnt"]==1
    assert ing["capabilities"]["resources"] is True
    assert ing["capabilities"]["calendars"] is True


def test_excel_progress_actual_dates_and_constraints_are_normalized():
    wb=Workbook()
    ws=wb.active
    ws.append(["Activity ID","Activity Name","Duration","Remaining Duration","% Complete","Actual Start","Actual Finish","Constraint Type","Constraint Date"])
    ws.append(["A100","Foundation","10 days","4 days",60,"2026-01-01","", "Must Finish By","2026-01-15"])
    ws.append(["A200","Completed Work","5 days","0 days",100,"2026-01-02","2026-01-06","",""])
    stream=io.BytesIO();wb.save(stream)

    ing=ingest_schedule("xlsx",stream.getvalue())
    tasks={x["task_code"]:x for x in ing["tables"]["TASK"]}
    assert tasks["A100"]["status_code"]=="In Progress"
    assert tasks["A100"]["act_start_date"]=="2026-01-01"
    assert tasks["A100"]["remain_drtn_hr_cnt"]==32
    assert tasks["A100"]["cstr_type"]=="Must Finish By"
    assert tasks["A100"]["cstr_date"]=="2026-01-15"
    assert tasks["A200"]["status_code"]=="Complete"
    assert tasks["A200"]["act_end_date"]=="2026-01-06"
