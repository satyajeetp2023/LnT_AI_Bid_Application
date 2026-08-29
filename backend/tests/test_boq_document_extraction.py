import io

from openpyxl import Workbook

from app.services.boq_document_extraction import extract_boq_rows


def test_excel_boq_table_is_detected_and_extracted():
    wb=Workbook()
    ws=wb.active
    ws.title="BOQ"
    ws.append(["Item No","Description of Work","Unit","Qty","Rate","Amount"])
    ws.append(["1","Design, supply and install OHE mast","No.",25,100,2500])
    ws.append(["2","Testing and commissioning of OHE section","Lot",1,500,500])
    stream=io.BytesIO();wb.save(stream)
    result=extract_boq_rows("xlsx",stream.getvalue())
    assert result["detected"] is True
    assert result["summary"]["rows"]==2
    assert result["rows"][0]["item_no"]=="BOQ:1"
    assert "OHE mast" in result["rows"][0]["description"]


def test_non_boq_excel_template_is_not_detected():
    wb=Workbook()
    ws=wb.active
    ws["A1"]="STATEMENT OF COMPLIANCE"
    ws["A2"]="Clause Reference"
    ws["B2"]="Yes"
    ws["C2"]="No"
    stream=io.BytesIO();wb.save(stream)
    result=extract_boq_rows("xlsx",stream.getvalue())
    assert result["detected"] is False
    assert result["rows"]==[]
