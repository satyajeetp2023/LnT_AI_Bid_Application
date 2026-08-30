import io

from openpyxl import Workbook

from app.services.template_structure_parser import parse_xlsx_template


def _sample_workbook():
    wb=Workbook()
    ws=wb.active
    ws.title="1.1"
    for index,name in enumerate(("1.1","1.1.1","1.1.1.1")):
        if index:
            ws=wb.create_sheet(name)
        ws.merge_cells("B1:F1")
        ws.merge_cells("B2:F2")
        ws.merge_cells("B3:F3")
        ws.merge_cells("B4:B5")
        ws.merge_cells("C4:D4")
        ws["B2"]="STATEMENT OF COMPLIANCE LIST"
        ws["B3"]="TENDERER---------------- DOCUMENT----------------"
        ws["B4"]="Clause Reference"
        ws["C4"]="Compliant"
        ws["E4"]="If not Compliant Tenderer's Comments/Proposal"
        ws["F4"]="Evaluator's Remarks"
        ws["C5"]="Yes"
        ws["D5"]="No"
        ws["B6"]=name
    out=io.BytesIO();wb.save(out);return out.getvalue()


def test_multirow_compliance_table_and_repeated_sheet_pattern():
    result=parse_xlsx_template(_sample_workbook())
    assert result["summary"]["compliance_tables"]==3
    assert result["summary"]["repeated_sheet_patterns"]==1
    first=result["tables"][0]
    assert first["table_type"]=="statement_of_compliance"
    assert first["header_rows"]==[4,5]
    semantics={c["semantic_field"] for c in first["columns"]}
    assert {"clause_reference","compliant_yes","compliant_no","tenderer_comments","evaluator_remarks"}<=semantics
    evaluator=next(c for c in first["columns"] if c["semantic_field"]=="evaluator_remarks")
    assert evaluator["ownership"]=="employer_only"
